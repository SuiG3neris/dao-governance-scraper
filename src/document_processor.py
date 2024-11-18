# src/document_processor.py

import os
import logging
from pathlib import Path
from typing import Dict, List, Optional, Union, Generator
from enum import Enum
from datetime import datetime
import mimetypes
import magic  # for better file type detection
import queue
from dataclasses import dataclass
from concurrent.futures import ThreadPoolExecutor

import pandas as pd
import PyPDF2
import pdfplumber
import pytesseract
from PIL import Image
from openpyxl import load_workbook
from docx import Document
from PyQt6.QtCore import QThread, pyqtSignal

# Custom exceptions
class DocumentProcessingError(Exception):
    """Base exception for document processing errors."""
    pass

class FileTypeError(DocumentProcessingError):
    """Raised when file type is unsupported or detection fails."""
    pass

class ExtractionError(DocumentProcessingError):
    """Raised when data extraction fails."""
    pass

@dataclass
class ProcessingResult:
    """Container for processing results."""
    success: bool
    file_path: Path
    file_type: str
    extracted_text: Optional[str] = None
    extracted_data: Optional[Dict] = None
    error_message: Optional[str] = None
    processing_time: float = 0.0
    metadata: Dict = None

class FileType(Enum):
    """Supported file types."""
    PDF = "pdf"
    IMAGE = "image"
    EXCEL = "excel"
    CSV = "csv"
    DOC = "doc"
    DOCX = "docx"
    TEXT = "text"
    UNKNOWN = "unknown"

class BaseProcessor:
    """Base class for file processors."""
    
    def __init__(self, config: Dict):
        self.config = config
        self.logger = logging.getLogger(self.__class__.__name__)

    def extract(self, file_path: Path) -> ProcessingResult:
        """Extract data from file."""
        raise NotImplementedError
        
    def validate(self, file_path: Path) -> bool:
        """Validate file before processing."""
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        return True

class PDFProcessor(BaseProcessor):
    """Process PDF files using both PyPDF2 and pdfplumber for robustness."""
    
    def extract(self, file_path: Path) -> ProcessingResult:
        self.validate(file_path)
        text = []
        try:
            # Try PyPDF2 first
            with open(file_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                for page in reader.pages:
                    text.append(page.extract_text())
                    
            # If text extraction was poor, try pdfplumber
            if not "".join(text).strip():
                with pdfplumber.open(file_path) as pdf:
                    text = [page.extract_text() for page in pdf.pages]
                    
            return ProcessingResult(
                success=True,
                file_path=file_path,
                file_type=FileType.PDF.value,
                extracted_text="\n".join(text),
                metadata={
                    'num_pages': len(text),
                    'file_size': file_path.stat().st_size
                }
            )
            
        except Exception as e:
            self.logger.error(f"PDF processing error: {str(e)}")
            return ProcessingResult(
                success=False,
                file_path=file_path,
                file_type=FileType.PDF.value,
                error_message=str(e)
            )

class ImageProcessor(BaseProcessor):
    """Process images using Tesseract OCR."""
    
    def extract(self, file_path: Path) -> ProcessingResult:
        self.validate(file_path)
        try:
            image = Image.open(file_path)
            text = pytesseract.image_to_string(image)
            
            return ProcessingResult(
                success=True,
                file_path=file_path,
                file_type=FileType.IMAGE.value,
                extracted_text=text,
                metadata={
                    'size': image.size,
                    'mode': image.mode,
                    'format': image.format
                }
            )
            
        except Exception as e:
            self.logger.error(f"Image processing error: {str(e)}")
            return ProcessingResult(
                success=False,
                file_path=file_path,
                file_type=FileType.IMAGE.value,
                error_message=str(e)
            )

class ExcelProcessor(BaseProcessor):
    """Process Excel files using openpyxl."""
    
    def extract(self, file_path: Path) -> ProcessingResult:
        self.validate(file_path)
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            data = {}
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                data[sheet_name] = []
                for row in sheet.rows:
                    data[sheet_name].append([cell.value for cell in row])
                    
            return ProcessingResult(
                success=True,
                file_path=file_path,
                file_type=FileType.EXCEL.value,
                extracted_data=data,
                metadata={
                    'sheets': wb.sheetnames,
                    'file_size': file_path.stat().st_size
                }
            )
            
        except Exception as e:
            self.logger.error(f"Excel processing error: {str(e)}")
            return ProcessingResult(
                success=False,
                file_path=file_path,
                file_type=FileType.EXCEL.value,
                error_message=str(e)
            )

class BatchProcessor(QThread):
    """Handle batch processing of multiple files."""
    
    progress = pyqtSignal(int)  # Progress percentage
    file_complete = pyqtSignal(str, ProcessingResult)  # File path and result
    batch_complete = pyqtSignal(list)  # List of all results
    error = pyqtSignal(str)  # Error message

    def __init__(self, config: Dict):
        super().__init__()
        self.config = config
        self.queue = queue.Queue()
        self.results = []
        self.running = True
        self.processors = {}  # File type to processor mapping
        self._init_processors()

    def _init_processors(self):
        """Initialize processor instances."""
        self.processors = {
            FileType.PDF: PDFProcessor(self.config),
            FileType.IMAGE: ImageProcessor(self.config),
            FileType.EXCEL: ExcelProcessor(self.config)
            # Add other processors as needed
        }

    def add_files(self, file_paths: List[Path]):
        """Add files to processing queue."""
        for path in file_paths:
            self.queue.put(path)

    def stop(self):
        """Stop processing."""
        self.running = False

    def run(self):
        """Process files in queue."""
        try:
            total_files = self.queue.qsize()
            processed = 0

            with ThreadPoolExecutor(max_workers=self.config.get('max_workers', 4)) as executor:
                while self.running and not self.queue.empty():
                    file_path = self.queue.get()
                    
                    try:
                        # Detect file type and get appropriate processor
                        file_type = self._detect_file_type(file_path)
                        processor = self.processors.get(file_type)
                        
                        if not processor:
                            raise FileTypeError(f"No processor for file type: {file_type}")
                            
                        # Process file
                        result = processor.extract(file_path)
                        self.results.append(result)
                        self.file_complete.emit(str(file_path), result)
                        
                        # Update progress
                        processed += 1
                        progress = int((processed / total_files) * 100)
                        self.progress.emit(progress)
                        
                    except Exception as e:
                        self.error.emit(f"Error processing {file_path}: {str(e)}")
                        
                    finally:
                        self.queue.task_done()

            self.batch_complete.emit(self.results)
            
        except Exception as e:
            self.error.emit(f"Batch processing error: {str(e)}")

    def _detect_file_type(self, file_path: Path) -> FileType:
        """Detect file type using multiple methods."""
        try:
            # Use python-magic for initial detection
            mime = magic.from_file(str(file_path), mime=True)
            
            # Map mime types to FileType enum
            mime_map = {
                'application/pdf': FileType.PDF,
                'image/': FileType.IMAGE,
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': FileType.EXCEL,
                'text/csv': FileType.CSV,
                'application/msword': FileType.DOC,
                'application/vnd.openxmlformats-officedocument.wordprocessingml.document': FileType.DOCX,
                'text/plain': FileType.TEXT
            }
            
            for mime_type, file_type in mime_map.items():
                if mime.startswith(mime_type):
                    return file_type
                    
            return FileType.UNKNOWN
            
        except Exception as e:
            logging.error(f"File type detection error: {str(e)}")
            return FileType.UNKNOWN